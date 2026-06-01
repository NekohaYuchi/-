import os
os.chdir(os.path.dirname(os.path.abspath(__file__)))
from datetime import datetime
import base64
import io

import pandas as pd
import numpy as np
import plotly.express as px
import joblib
from sklearn.metrics import confusion_matrix, roc_auc_score
import dash
from dash import dcc, html, dash_table, Input, Output, State, ctx, no_update
import dash_bootstrap_components as dbc

# ── 配色常數 ──────────────────────────────────────────────
COLORS = {
    "bg_page":    "#0B111A",
    "bg_card":    "#1F2937",
    "bg_sidebar": "#111827",
    "bg_input":   "#273142",
    "accent":     "#1E9FFB",
    "cyan":       "#22D3EE",
    "text":       "#F5F7FA",
    "text_muted": "#9CA3AF",
    "fraud":      "#EF4444",
    "safe":       "#00D084",
    "warning":    "#F59E0B",
    "border":     "#2F3B4A",
}

# ── 載入模型與測試資料 ─────────────────────────────────────
print("載入模型與測試數據...")
model_pipeline = joblib.load("fraud_model_pipeline.pkl")
X_test, y_test = joblib.load("test_data.pkl")
scaler        = joblib.load("scaler.pkl")
print("載入成功")

y_proba_all = model_pipeline.predict_proba(X_test)[:, 1]
y_all = y_test.values

# 固定 AUC（只算一次）
AUC_SCORE = round(float(roc_auc_score(y_test, y_proba_all)), 3)

# 預先計算 99 個閥值的混淆矩陣統計
thresholds_array = np.arange(0.01, 1.00, 0.01)
tp_arr = np.zeros_like(thresholds_array)
fp_arr = np.zeros_like(thresholds_array)
fn_arr = np.zeros_like(thresholds_array)
tn_arr = np.zeros_like(thresholds_array)

for i, t in enumerate(thresholds_array):
    y_pred_t = (y_proba_all >= t).astype(int)
    cm_t = confusion_matrix(y_test, y_pred_t)
    if cm_t.size == 4:
        tn_arr[i], fp_arr[i], fn_arr[i], tp_arr[i] = cm_t.ravel()
    else:
        tn_arr[i] = len(y_test) - sum(y_pred_t)
        tp_arr[i] = sum(y_pred_t)

CURVE_STATS = pd.DataFrame({
    "Threshold": thresholds_array,
    "TP": tp_arr, "FP": fp_arr,
    "FN": fn_arr, "TN": tn_arr,
})

# 預先繪製特徵重要性圖（固定，不在 callback 中重繪）
xgb_clf = model_pipeline.named_steps["classifier"]
importances = xgb_clf.feature_importances_
indices = np.argsort(importances)[::-1][:10]
top_features = X_test.columns[indices]
top_importances = importances[indices]

FIG_IMPORTANCE = px.bar(
    x=top_importances, y=top_features, orientation="h",
    title="XGBoost 全局核心風控特徵 (Top 10)",
    labels={"x": "相對重要性得分", "y": "加密特徵欄位 (PCA)"},
    color=top_importances,
    color_continuous_scale=[[0, "#1a3a5c"], [0.5, "#1E9FFB"], [1, "#22D3EE"]],
)
FIG_IMPORTANCE.update_layout(
    yaxis={"categoryorder": "total ascending",
           "gridcolor": COLORS["border"], "gridwidth": 1},
    xaxis={"gridcolor": COLORS["border"], "gridwidth": 1},
    paper_bgcolor=COLORS["bg_card"],
    plot_bgcolor=COLORS["bg_card"],
    font={"color": COLORS["text"]},
    margin={"l": 10, "r": 10, "b": 30, "t": 40},
    coloraxis_showscale=False,
)

# 每日報表交易池（server-side 全域，15k 筆索引不適合放 dcc.Store）
WINDOW_SIZE = 15786
recent_tx_indices = list(
    np.random.choice(np.arange(len(y_test)), WINDOW_SIZE, replace=False)
)

# ── 版面 helpers ──────────────────────────────────────────

def _metric_row(label, value, value_color=None):
    return html.Div([
        html.Span(label, style={"color": COLORS["text_muted"], "fontSize": "11px",
                                 "width": "120px", "display": "inline-block"}),
        html.Span(value, style={"color": value_color or COLORS["text"],
                                 "fontFamily": "monospace", "fontSize": "12px"}),
    ], style={"marginBottom": "5px"})


def make_kpi_card(title, value_id, color):
    return dbc.Card([
        dbc.CardBody([
            html.P(title, style={"color": COLORS["text_muted"], "fontSize": "11px",
                                  "marginBottom": "4px"}),
            html.H4(id=value_id, children="—",
                    style={"color": color, "fontWeight": "bold", "marginBottom": 0}),
        ], style={"padding": "12px"})
    ], style={"backgroundColor": COLORS["bg_card"],
              "border": f"1px solid {COLORS['border']}"})


EMPTY_TABLE_ROWS = [
    {"交易ID": "-", "風險評分": "-", "系統判定狀態": "-", "處置優先級": "-"}
] * 5

TABLE_KWARGS = dict(
    style_table={"overflowX": "auto", "overflowY": "auto", "maxHeight": "320px"},
    style_header={
        "backgroundColor": COLORS["bg_sidebar"],
        "color": COLORS["cyan"], "fontWeight": "600",
        "border": f"1px solid {COLORS['border']}",
        "fontSize": "11px", "letterSpacing": "0.05em", "textTransform": "uppercase",
    },
    style_cell={
        "backgroundColor": COLORS["bg_input"],
        "color": COLORS["text_muted"],
        "textAlign": "center", "padding": "9px 8px",
        "border": f"1px solid {COLORS['border']}",
        "fontFamily": "monospace", "fontSize": "13px",
    },
    style_data_conditional=[
        {"if": {"filter_query": '{系統判定狀態} contains "詐騙"'},
         "color": COLORS["fraud"], "fontWeight": "bold"},
        {"if": {"filter_query": '{系統判定狀態} contains "安全"'},
         "color": COLORS["safe"]},
    ],
)


def page_monitor():
    return html.Div([
        dbc.Row([
            dbc.Col([
                html.P(id="status-display", children="—",
                       style={"fontSize": "22px", "fontWeight": "bold",
                               "color": COLORS["text"], "textAlign": "center",
                               "margin": "8px 0"}),
                html.Pre(id="sop-display", children="—",
                         style={"color": COLORS["text_muted"], "fontSize": "12px",
                                "whiteSpace": "pre-wrap", "marginBottom": 0}),
            ], width=12),
        ], className="mb-3"),
        dbc.Row([
            dbc.Col([
                html.H6("銀行後台實時交易風險監控流",
                        style={"color": COLORS["text"], "marginBottom": "8px"}),
                dash_table.DataTable(
                    id="stream-table",
                    columns=[{"name": c, "id": c} for c in
                             ["交易ID", "風險評分", "系統判定狀態", "處置優先級"]],
                    data=EMPTY_TABLE_ROWS,
                    row_selectable="single",
                    selected_rows=[],
                    **TABLE_KWARGS,
                ),
                html.Div(id="tx-detail-panel", style={"marginTop": "12px"}),
            ], width=8),
            dbc.Col([
                html.Pre(id="profit-display", children="—",
                         style={"color": COLORS["text"], "fontSize": "12px",
                                "whiteSpace": "pre-wrap",
                                "backgroundColor": COLORS["bg_card"],
                                "padding": "12px", "borderRadius": "6px",
                                "border": f"1px solid {COLORS['border']}",
                                "height": "100%"}),
            ], width=4),
        ]),
    ])


def page_analysis():
    return html.Div([
        dbc.Row([
            dbc.Col([
                dbc.Card(dbc.CardBody(
                    dcc.Graph(id="importance-chart", figure=FIG_IMPORTANCE,
                              config={"displayModeBar": False})
                ), style={"backgroundColor": COLORS["bg_card"],
                           "border": f"1px solid {COLORS['border']}"}),
            ], width=6),
            dbc.Col([
                dbc.Card(dbc.CardBody(
                    dcc.Graph(id="profit-curve", config={"displayModeBar": False})
                ), style={"backgroundColor": COLORS["bg_card"],
                           "border": f"1px solid {COLORS['border']}"}),
            ], width=6),
        ]),
    ])


def page_report():
    return html.Div([
        dbc.Row([
            dbc.Col([
                dbc.Button("產生管理報表", id="report-btn", color="primary", className="mb-3"),
                html.Div(id="report-content-area"),
                dcc.Download(id="report-download"),
                dbc.Button("下載 Excel 報表", id="download-btn",
                           color="success", style={"display": "none"}, className="mt-3"),
            ]),
        ]),
    ])


# ── Dash app ──────────────────────────────────────────────
app = dash.Dash(
    __name__,
    external_stylesheets=[dbc.themes.SLATE],
    suppress_callback_exceptions=True,
    title="銀行智慧型信用卡防詐系統",
)

_sidebar_style = {
    "position": "fixed", "top": 0, "left": 0, "bottom": 0,
    "width": "260px", "padding": "16px",
    "backgroundColor": COLORS["bg_sidebar"],
    "borderRight": f"1px solid {COLORS['border']}",
    "overflowY": "auto",
    "zIndex": 100,
}
_content_style = {
    "marginLeft": "276px",
    "padding": "20px",
    "backgroundColor": COLORS["bg_page"],
    "minHeight": "100vh",
}

app.layout = html.Div([
    dcc.Location(id="url"),
    dcc.Store(id="tx-state",       data={"id": None, "score": 0.0}),
    dcc.Store(id="history-store",  data=[]),
    dcc.Store(id="settings-store", data={"threshold": 0.50, "cost_fn": 250, "cost_fp": 15}),
    dcc.Store(id="report-store"),
    dcc.Store(id="session-stats", data={"intercepted": 0, "protected": 0.0}),
    dcc.Interval(id="init-interval",   interval=100,  max_intervals=1),
    dcc.Interval(id="stream-interval", interval=500, disabled=True),

    # ── 側邊欄 ──────────────────────────────────────
    html.Div([
        html.H6("信用卡防詐系統",
                style={"color": COLORS["text"], "fontSize": "14px", "marginBottom": "2px"}),
        html.P("工資系 商管程式設計 期末專題",
               style={"color": COLORS["text_muted"], "fontSize": "11px", "marginBottom": "12px"}),
        html.Hr(style={"borderColor": COLORS["border"]}),

        dbc.Nav([
            dbc.NavLink("監控", href="/",         active="exact",
                        style={"color": COLORS["text"], "borderRadius": "6px",
                               "marginBottom": "4px", "fontSize": "13px"}),
            dbc.NavLink("分析", href="/analysis",  active="exact",
                        style={"color": COLORS["text"], "borderRadius": "6px",
                               "marginBottom": "4px", "fontSize": "13px"}),
            dbc.NavLink("報表", href="/report",    active="exact",
                        style={"color": COLORS["text"], "borderRadius": "6px",
                               "marginBottom": "4px", "fontSize": "13px"}),
        ], vertical=True, pills=True),

        html.Hr(style={"borderColor": COLORS["border"]}),

        html.Label("銀行營運方針", style={"color": COLORS["text"], "fontSize": "12px"}),
        dcc.Dropdown(
            id="scenario-dropdown",
            options=[
                {"label": "損益平衡最優化", "value": "balanced"},
                {"label": "嚴格資安防禦",   "value": "strict"},
                {"label": "客戶體驗優先",   "value": "ux"},
            ],
            value="balanced", clearable=False,
            style={"fontSize": "12px", "marginBottom": "12px"},
        ),
        html.Label("風險判定閥值", style={"color": COLORS["text"], "fontSize": "12px"}),
        dcc.Slider(id="threshold-slider", min=0.01, max=0.99, step=0.01, value=0.50,
                   tooltip={"placement": "bottom", "always_visible": True},
                   marks={0.01: "0", 0.5: "0.5", 0.99: "1"},
                   className="mb-3"),
        html.Label("漏抓成本 FN Cost (USD/筆)", style={"color": COLORS["text"], "fontSize": "12px"}),
        dcc.Slider(id="cost-fn-slider", min=50, max=1000, step=10, value=250,
                   tooltip={"placement": "bottom", "always_visible": True},
                   marks={50: "50", 500: "500", 1000: "1000"},
                   className="mb-3"),
        html.Label("誤判成本 FP Cost (USD/筆)", style={"color": COLORS["text"], "fontSize": "12px"}),
        dcc.Slider(id="cost-fp-slider", min=1, max=100, step=1, value=15,
                   tooltip={"placement": "bottom", "always_visible": True},
                   marks={1: "1", 50: "50", 100: "100"},
                   className="mb-3"),
        dbc.Button("接收即時交易", id="simulate-btn", color="primary",
                   size="sm", className="w-100 mt-2"),
    ], style=_sidebar_style),

    # ── 主內容區 ──────────────────────────────────────
    html.Div([
        # KPI 卡（所有頁面皆顯示）
        dbc.Row([
            dbc.Col(make_kpi_card("今日攔截筆數",    "kpi-intercepted", COLORS["fraud"]),   width=3),
            dbc.Col(make_kpi_card("防護金額 (USD)",   "kpi-saved",       COLORS["safe"]),    width=3),
            dbc.Col(make_kpi_card("模型 AUC",          "kpi-auc",         COLORS["cyan"]),    width=3),
            dbc.Col(make_kpi_card("當前閥值",           "kpi-threshold",   COLORS["warning"]), width=3),
        ], className="mb-3 g-2"),

        # 三個頁面（全在 DOM，CSS 切換顯示）
        html.Div(id="page-monitor",  children=page_monitor()),
        html.Div(id="page-analysis", children=page_analysis(), style={"display": "none"}),
        html.Div(id="page-report",   children=page_report(),   style={"display": "none"}),
    ], style=_content_style),
], style={"backgroundColor": COLORS["bg_page"]})

# ── Callbacks ─────────────────────────────────────────────

@app.callback(
    Output("page-monitor",  "style"),
    Output("page-analysis", "style"),
    Output("page-report",   "style"),
    Input("url", "pathname"),
)
def route_page(pathname):
    show = {"display": "block"}
    hide = {"display": "none"}
    if pathname == "/analysis":
        return hide, show, hide
    if pathname == "/report":
        return hide, hide, show
    return show, hide, hide


@app.callback(
    Output("threshold-slider", "value"),
    Input("scenario-dropdown", "value"),
)
def handle_scenario(scenario):
    return {"strict": 0.20, "ux": 0.80, "balanced": 0.50}.get(scenario, 0.50)


@app.callback(
    Output("stream-interval", "disabled"),
    Output("simulate-btn",    "children"),
    Output("simulate-btn",    "color"),
    Input("simulate-btn",     "n_clicks"),
    State("stream-interval",  "disabled"),
    prevent_initial_call=True,
)
def toggle_stream(n_clicks, is_disabled):
    if is_disabled:
        return False, "停止接收", "danger"
    return True, "接收即時交易", "primary"


@app.callback(
    Output("kpi-intercepted",  "children"),
    Output("kpi-saved",        "children"),
    Output("kpi-auc",          "children"),
    Output("kpi-threshold",    "children"),
    Output("status-display",   "children"),
    Output("sop-display",      "children"),
    Output("profit-display",   "children"),
    Output("stream-table",     "data"),
    Output("stream-table",     "selected_rows"),
    Output("profit-curve",     "figure"),
    Output("tx-state",         "data"),
    Output("history-store",    "data"),
    Output("settings-store",   "data"),
    Output("session-stats",    "data"),
    Input("init-interval",     "n_intervals"),
    Input("stream-interval",   "n_intervals"),
    Input("simulate-btn",      "n_clicks"),
    Input("threshold-slider",  "value"),
    Input("cost-fn-slider",    "value"),
    Input("cost-fp-slider",    "value"),
    State("tx-state",          "data"),
    State("history-store",     "data"),
    State("session-stats",     "data"),
    prevent_initial_call=False,
)
def update_dashboard(n_intervals, n_stream, n_clicks, threshold, cost_fn, cost_fp,
                     tx_state, history_data, session_stats):
    triggered = ctx.triggered_id
    is_new_tx = triggered in ("simulate-btn", "init-interval", "stream-interval") or tx_state["id"] is None

    # ── 抽樣一筆交易 ─────────────────────────────────
    if is_new_tx:
        if np.random.rand() < 0.30:
            idx = int(np.random.choice(np.where(y_all == 1)[0]))
        else:
            idx = int(np.random.choice(np.where(y_all == 0)[0]))
        sample = X_test.iloc[[idx]]
        risk   = float(model_pipeline.predict_proba(sample)[0][1])
        tx_id  = f"#TX-{np.random.randint(10000, 99999)}"

        # 反標準化取得原始金額與時間
        time_std   = float(sample["Time"].values[0])
        amount_std = float(sample["Amount"].values[0])
        original   = scaler.inverse_transform([[time_std, amount_std]])[0]
        time_sec   = float(original[0])
        amount_eur = float(original[1])
        true_label = int(y_all[idx])

        tx_state = {
            "id": tx_id, "score": risk,
            "time_sec": time_sec, "amount_eur": amount_eur,
            "true_label": true_label,
        }
    else:
        tx_id      = tx_state["id"]
        risk       = tx_state["score"]
        time_sec   = tx_state.get("time_sec", 0.0)
        amount_eur = tx_state.get("amount_eur", 0.0)
        true_label = tx_state.get("true_label", -1)

    # ── 判定結果與 SOP ───────────────────────────────
    is_fraud   = risk >= threshold
    status_txt = "[!] 詐騙警示 (FRAUD)" if is_fraud else "[OK] 安全交易 (SAFE)"
    if is_fraud:
        sop = (
            "[風控系統核心響應機制觸發]:\n"
            "1. 該筆交易已被即時攔截，暫停授權發放。\n"
            "2. 系統已自動透過 LINE/簡訊 發送消費確認通知給持卡人。\n"
            "3. 風控中心已自動將此案派發工單至『一線客服部』進行人工外撥核對。"
        )
    else:
        sop = (
            "[標準商務流程放行]:\n"
            "1. 交易通過 XGBoost 演算法風控比對，毫秒級授權成功。\n"
            "2. 計入當日持卡人正常消費信用額度，保障刷卡順暢體驗。"
        )

    priority = (
        "Level 1 (緊急阻斷)" if risk > 0.90 else
        "Level 2 (人工確認)" if is_fraud else
        "Level 3 (自動化驗證)"
    )

    # ── 累積今日攔截統計 ─────────────────────────────
    if is_new_tx and is_fraud:
        session_stats = {
            "intercepted": session_stats["intercepted"] + 1,
            "protected":   session_stats["protected"]   + max(amount_eur, 0.0),
        }

    # ── 更新交易歷史 ─────────────────────────────────
    if is_new_tx:
        history_data = [{
            "交易ID": tx_id, "風險評分": round(risk, 4),
            "系統判定狀態": status_txt, "處置優先級": priority,
            "_risk_raw":   risk,       "_is_fraud":   is_fraud,
            "_sop":        sop,        "_true_label": true_label,
            "_amount_eur": amount_eur, "_time_sec":   time_sec,
        }] + history_data
        history_data = history_data[:20]
    elif history_data:
        history_data[0]["系統判定狀態"] = status_txt
        history_data[0]["處置優先級"]   = priority

    table_rows = history_data if history_data else EMPTY_TABLE_ROWS

    # ── 利潤計算 ─────────────────────────────────────
    y_pred = (y_proba_all >= threshold).astype(int)
    cm = confusion_matrix(y_test, y_pred)
    if cm.size == 4:
        tn, fp, fn, tp = cm.ravel()
    else:
        tn = int(len(y_test) - sum(y_pred))
        fp = fn = 0
        tp = int(sum(y_pred))

    prevented = int(tp) * cost_fn
    friction  = int(fp) * cost_fp + int(fn) * cost_fn
    net       = prevented - friction

    profit_txt = (
        f"全局財務效益試算:\n"
        f"- 成功防堵呆帳: +${prevented:,} USD (攔截 {tp} 筆)\n"
        f"- 營收摩擦與漏抓成本: -${friction:,} USD\n"
        f"銀行風控純防護收益淨值: ${net:,} USD"
    )

    # ── 利潤曲線 ─────────────────────────────────────
    df_curve = CURVE_STATS.copy()
    df_curve["Profit"] = (df_curve["TP"] * cost_fn
                          - (df_curve["FP"] * cost_fp + df_curve["FN"] * cost_fn))
    fig_curve = px.line(
        df_curve, x="Threshold", y="Profit",
        title="動態利潤最佳化曲線圖",
        labels={"Threshold": "判定閥值", "Profit": "預期淨利潤 (USD)"},
        color_discrete_sequence=[COLORS["accent"]],
    )
    fig_curve.add_scatter(
        x=[threshold], y=[net], mode="markers",
        marker={"size": 10, "color": COLORS["warning"], "symbol": "diamond"},
        name="當前決策點",
    )
    min_p, max_p = df_curve["Profit"].min(), df_curve["Profit"].max()
    p_range = max(float(max_p - min_p), 1.0)
    fig_curve.update_layout(
        xaxis={"range": [0, 1], "gridcolor": COLORS["border"], "gridwidth": 1},
        yaxis={"range": [min_p - p_range * 0.05, max_p + p_range * 0.05],
               "gridcolor": COLORS["border"], "gridwidth": 1},
        paper_bgcolor=COLORS["bg_card"], plot_bgcolor=COLORS["bg_card"],
        font={"color": COLORS["text"]},
        margin={"l": 10, "r": 10, "b": 30, "t": 40},
        showlegend=False,
    )

    settings = {"threshold": threshold, "cost_fn": cost_fn, "cost_fp": cost_fp}

    return (
        str(session_stats["intercepted"]),
        f"${session_stats['protected']:,.2f}",
        str(AUC_SCORE), str(round(threshold, 2)),
        status_txt, sop, profit_txt, table_rows, [],
        fig_curve, tx_state, history_data, settings, session_stats,
    )


@app.callback(
    Output("report-content-area", "children"),
    Output("report-store",        "data"),
    Output("download-btn",        "style"),
    Input("report-btn",           "n_clicks"),
    State("settings-store",       "data"),
    prevent_initial_call=True,
)
def generate_report(n_clicks, settings):
    threshold = settings["threshold"]
    cost_fn   = settings["cost_fn"]
    cost_fp   = settings["cost_fp"]

    # ── 基礎統計 ──────────────────────────────────────
    recent_y     = y_test.iloc[recent_tx_indices]
    recent_proba = y_proba_all[recent_tx_indices]
    y_pred       = (recent_proba >= threshold).astype(int)
    labels       = recent_y.values

    cm = confusion_matrix(recent_y, y_pred, labels=[0, 1])
    tn, fp, fn, tp = cm.ravel() if cm.size == 4 else (
        int(len(recent_y) - sum(y_pred)), 0, 0, int(sum(y_pred)))

    # ── 原始金額與時間（inverse transform） ────────────
    orig = scaler.inverse_transform(
        X_test.iloc[recent_tx_indices][['Time', 'Amount']])
    time_arr   = orig[:, 0]
    amount_arr = orig[:, 1]

    # ── 財務計算 ──────────────────────────────────────
    saved     = int(tp) * cost_fn
    fp_cost   = int(fp) * cost_fp
    fn_cost   = int(fn) * cost_fn
    friction  = fp_cost + fn_cost
    net       = saved - friction

    # ── 模型指標 ──────────────────────────────────────
    precision = tp / (tp + fp)  if (tp + fp) > 0  else 0.0
    recall    = tp / (tp + fn)  if (tp + fn) > 0  else 0.0
    f1        = 2*precision*recall / (precision+recall) if (precision+recall) > 0 else 0.0
    profit_arr  = (CURVE_STATS["TP"] * cost_fn
                   - (CURVE_STATS["FP"] * cost_fp + CURVE_STATS["FN"] * cost_fn))
    best_thresh = float(CURVE_STATS["Threshold"].iloc[np.argmax(profit_arr)])
    report_time = datetime.now().strftime("%Y-%m-%d %H:%M")

    # ── 高風險交易清單（FP + 高風險 FN） ──────────────
    fp_mask = (y_pred == 1) & (labels == 0)
    fn_mask = (y_pred == 0) & (labels == 1) & (recent_proba > 0.3)
    hr_idx_all = np.where(fp_mask | fn_mask)[0]
    hr_idx_all = hr_idx_all[np.argsort(-recent_proba[hr_idx_all])]  # 全部，依風險排序
    hr_idx_top = hr_idx_all[:20]                                      # 前 20 筆供 UI 顯示

    def fmt_time(sec):
        return f"T+{int(sec//3600)}h{int((sec%3600)//60):02d}m"

    def make_hr_row(i):
        return {
            "交易ID":   f"TX-{np.random.randint(100000,999999)}",
            "時間":     fmt_time(time_arr[i]),
            "金額":     f"${amount_arr[i]:,.2f}",
            "風險評分": round(float(recent_proba[i]), 4),
            "類型":     "FP — 誤報" if fp_mask[i] else "FN — 漏抓",
            "建議行動": "需主動回電解除凍結" if fp_mask[i] else "需人工複核",
        }

    highrisk_rows     = [make_hr_row(i) for i in hr_idx_top]   # UI 用（top 20）
    highrisk_rows_all = [make_hr_row(i) for i in hr_idx_all]   # CSV 用（全部）

    # ── Section helpers ──────────────────────────────
    def section_title(text):
        return html.P(text, style={
            "color": COLORS["cyan"], "fontSize": "12px", "fontWeight": "bold",
            "borderBottom": f"1px solid {COLORS['border']}",
            "paddingBottom": "4px", "marginBottom": "10px", "marginTop": "16px",
        })

    def stat_card(label, value, color=COLORS["text"]):
        return dbc.Col(dbc.Card(dbc.CardBody([
            html.P(label, style={"color": COLORS["text_muted"], "fontSize": "10px",
                                  "marginBottom": "2px"}),
            html.H5(value, style={"color": color, "fontFamily": "monospace",
                                   "fontWeight": "bold", "marginBottom": 0}),
        ], style={"padding": "10px"}), style={
            "backgroundColor": COLORS["bg_input"],
            "border": f"1px solid {COLORS['border']}"}), width=3)

    def cm_cell(value, bg, color):
        return html.Td(str(value), style={
            "backgroundColor": bg, "color": color,
            "fontFamily": "monospace", "textAlign": "center",
            "padding": "8px 16px", "fontWeight": "bold", "fontSize": "14px",
            "border": f"1px solid {COLORS['border']}",
        })

    # ── 組裝 HTML 報表 ────────────────────────────────
    content = html.Div([

        # Header
        html.Div([
            html.Span("風控管理報表", style={"color": COLORS["text"], "fontSize": "15px",
                                              "fontWeight": "bold"}),
            html.Span(f"  {report_time}", style={"color": COLORS["text_muted"],
                                                   "fontFamily": "monospace", "fontSize": "11px",
                                                   "marginLeft": "12px"}),
        ], style={"marginBottom": "4px"}),
        html.Span(
            f"Threshold: {threshold}  |  FN Cost: ${cost_fn}  |  FP Cost: ${cost_fp}",
            style={"color": COLORS["text_muted"], "fontSize": "11px", "fontFamily": "monospace"}),

        # 一、執行摘要
        section_title("一、執行摘要"),
        dbc.Row([
            stat_card("總交易量",   f"{len(recent_y):,} 筆"),
            stat_card("詐騙攔截率", f"{recall*100:.1f}%",   COLORS["safe"]),
            stat_card("防護金額",   f"${saved:,}",          COLORS["safe"]),
            stat_card("風控淨效益", f"${net:,}",
                      COLORS["safe"] if net >= 0 else COLORS["fraud"]),
        ], className="g-2 mb-2"),

        # 二、風控績效指標
        section_title("二、風控績效指標"),
        dbc.Row([
            dbc.Col([
                html.Table([
                    html.Tbody([
                        html.Tr([
                            html.Td(""),
                            html.Td("預測詐騙", style={"color": COLORS["fraud"],
                                                        "fontSize": "11px", "textAlign": "center",
                                                        "padding": "4px 12px"}),
                            html.Td("預測正常", style={"color": COLORS["safe"],
                                                        "fontSize": "11px", "textAlign": "center",
                                                        "padding": "4px 12px"}),
                        ]),
                        html.Tr([
                            html.Td("實際詐騙", style={"color": COLORS["fraud"],
                                                        "fontSize": "11px", "padding": "4px 8px"}),
                            cm_cell(tp, "#0d3320", COLORS["safe"]),
                            cm_cell(fn, "#3a1a1a", COLORS["fraud"]),
                        ]),
                        html.Tr([
                            html.Td("實際正常", style={"color": COLORS["safe"],
                                                        "fontSize": "11px", "padding": "4px 8px"}),
                            cm_cell(fp, "#2d2510", COLORS["warning"]),
                            cm_cell(f"{tn:,}", "#0d1a2e", COLORS["accent"]),
                        ]),
                    ])
                ], style={"borderCollapse": "collapse"}),
            ], width=5),
            dbc.Col([
                _metric_row("Precision",   f"{precision:.3f}  ({precision*100:.1f}%)"),
                _metric_row("Recall",      f"{recall:.3f}  ({recall*100:.1f}%)"),
                _metric_row("F1 Score",    f"{f1:.3f}"),
                _metric_row("Model AUC",   f"{AUC_SCORE}"),
                html.Hr(style={"borderColor": COLORS["border"], "margin": "6px 0"}),
                _metric_row("誤攔成本",    f"${fp_cost:,}  ({fp} 筆 × ${cost_fp})"),
                _metric_row("漏抓損失",    f"${fn_cost:,}  ({fn} 筆 × ${cost_fn})"),
                _metric_row("最佳閥值建議", f"{best_thresh:.2f}",    COLORS["warning"]),
            ], width=7),
        ], className="mb-2"),

        # 三、高風險交易清單
        section_title(f"三、高風險交易清單（FP + 高風險 FN，共 {len(highrisk_rows_all)} 筆，"
                       "介面顯示前 20 筆，完整清單請下載 Excel）"),
        dash_table.DataTable(
            columns=[{"name": c, "id": c} for c in
                     ["交易ID", "時間", "金額", "風險評分", "類型", "建議行動"]],
            data=highrisk_rows or [{"交易ID": "（本次無高風險交易）", "時間": "-",
                                     "金額": "-", "風險評分": "-",
                                     "類型": "-", "建議行動": "-"}],
            style_table={"overflowX": "auto"},
            style_header={"backgroundColor": COLORS["bg_sidebar"], "color": COLORS["cyan"],
                          "fontWeight": "600", "border": f"1px solid {COLORS['border']}",
                          "fontSize": "11px", "letterSpacing": "0.04em"},
            style_cell={"backgroundColor": COLORS["bg_input"], "color": COLORS["text_muted"],
                        "textAlign": "center", "padding": "7px 10px",
                        "border": f"1px solid {COLORS['border']}",
                        "fontFamily": "monospace", "fontSize": "12px"},
            style_data_conditional=[
                {"if": {"filter_query": '{類型} contains "FP"'},
                 "color": COLORS["warning"], "fontWeight": "bold"},
                {"if": {"filter_query": '{類型} contains "FN"'},
                 "color": COLORS["fraud"], "fontWeight": "bold"},
            ],
        ),
    ])

    # ── Excel 報告 ──
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    summary_rows = [
        ["總交易量", len(recent_y), "筆"],
        ["詐騙攔截率", recall, ""],
        ["防護金額", saved, "USD"],
        ["風控淨效益", net, "USD"],
    ]
    performance_rows = [
        ["TP（成功攔截）", tp, "真詐騙且被攔截"],
        ["FP（誤報）", fp, "實際正常但被攔截"],
        ["FN（漏抓）", fn, "真詐騙但未攔截"],
        ["TN（正確放行）", tn, "實際正常且放行"],
        ["Precision", precision, f"{precision*100:.1f}%"],
        ["Recall", recall, f"{recall*100:.1f}%"],
        ["F1 Score", f1, ""],
        ["Model AUC", AUC_SCORE, ""],
        ["誤攔成本", fp_cost, f"{fp} 筆 x ${cost_fp}"],
        ["漏抓損失", fn_cost, f"{fn} 筆 x ${cost_fn}"],
        ["最佳閥值建議", best_thresh, "以當前成本參數最大化淨效益"],
    ]
    highrisk_export_rows = [
        [row["交易ID"], row["時間"],
         float(row["金額"].replace("$", "").replace(",", "")),  # 數字而非字串，Excel 可排序
         row["風險評分"], row["類型"], row["建議行動"]]
        for row in highrisk_rows_all
    ]
    detail_rows = []
    for i in range(len(recent_y)):
        r, p = int(labels[i]), int(y_pred[i])
        result = ("TP" if r==1 and p==1 else "FP" if r==0 and p==1
                  else "FN" if r==1 and p==0 else "TN")
        prob = float(recent_proba[i])
        detail_rows.append([
            f"TX-{np.random.randint(100000,999999)}",
            fmt_time(time_arr[i]),
            round(float(amount_arr[i]), 2),
            round(prob, 4),
            "高" if prob > 0.7 else "中" if prob > 0.3 else "低",
            "詐騙" if r == 1 else "正常",
            "攔截" if p == 1 else "放行",
            result,
            ("Level 1 (緊急阻斷)" if prob > 0.9 and p == 1
             else "Level 2 (人工確認)" if p == 1
             else "Level 3 (自動化驗證)"),
        ])

    from openpyxl.formatting.rule import DataBarRule, FormulaRule

    # ─ 色票與工具 ─
    def _f(h): return PatternFill("solid", fgColor=h)
    def _s(c="CBD5E1"): return Side(style="thin", color=c)
    def _b4(c="CBD5E1"):
        s = _s(c); return Border(left=s, right=s, top=s, bottom=s)

    F_BANNER  = _f("1E3A8A"); F_PARAMS  = _f("EFF6FF")
    F_SECTION = _f("1D4ED8"); F_HDR     = _f("1E293B")
    F_ALT     = _f("F8FAFC"); F_WHITE   = _f("FFFFFF")
    F_GREEN   = _f("DCFCE7"); F_RED     = _f("FEE2E2"); F_ORANGE = _f("FEF3C7")
    FULL_B    = _b4();         BOT_B    = Border(bottom=_s())

    def _w(v):
        return sum(2 if ord(c) > 127 else 1 for c in str(v if v is not None else ""))

    def auto_w(sheet, pad=3, lo=12, hi=55):
        for col in sheet.columns:
            sheet.column_dimensions[get_column_letter(col[0].column)].width = \
                min(max(max(_w(c.value) for c in col) + pad, lo), hi)

    # ══ Sheet 1：管理摘要 ════════════════════════════════
    wb = Workbook()
    ws = wb.active
    ws.title = "管理摘要"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "1E3A8A"
    ws.freeze_panes = "A6"

    ws.append(["風控管理報表", "", "", "", report_time, ""])    # 1
    ws.append(["", f"Threshold = {threshold}",
               f"FN Cost = ${cost_fn}", f"FP Cost = ${cost_fp}", "", ""])  # 2
    ws.append([])                            # 3
    ws.append(["一、執行摘要"])               # 4
    ws.append(["指標", "數值", "單位"])       # 5
    for row in summary_rows:   ws.append(row)      # 6–9
    ws.append([])                            # 10
    ws.append(["二、風控績效指標"])            # 11
    ws.append(["指標", "數值", "備註"])       # 12
    for row in performance_rows: ws.append(row)    # 13–23

    # 橫幅 (row 1)
    ws.merge_cells("A1:D1")
    ws["A1"].fill = F_BANNER
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=15)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws["E1"].fill = F_BANNER
    ws["E1"].font = Font(color="BFD7FF", size=10, italic=True)
    ws["E1"].alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[1].height = 34

    # 參數列 (row 2)
    for col in range(1, 5):
        c = ws.cell(row=2, column=col)
        c.fill = F_PARAMS
        c.font = Font(color="1E40AF", size=10)
        c.alignment = Alignment(horizontal="left" if col == 1 else "center", indent=1)
        c.border = BOT_B
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 5     # 細分隔

    # 執行摘要 section 標題 (row 4)
    ws.merge_cells("A4:C4")
    ws["A4"].fill = F_SECTION
    ws["A4"].font = Font(color="FFFFFF", bold=True, size=11)
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 22

    # 欄位標題 (row 5)
    for col in range(1, 4):
        c = ws.cell(row=5, column=col)
        c.fill = F_HDR; c.border = FULL_B
        c.font = Font(color="FFFFFF", bold=True, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[5].height = 20

    # 摘要資料 (rows 6–9)，交替底色
    for r in range(6, 10):
        f = F_ALT if r % 2 == 0 else F_WHITE
        ws.row_dimensions[r].height = 20
        for col, (ha, fnt) in enumerate([
            ("left",  Font(size=10)),
            ("right", Font(bold=True, size=11)),
            ("left",  Font(color="6B7280", size=9, italic=True)),
        ], 1):
            c = ws.cell(row=r, column=col)
            c.fill = f; c.border = FULL_B
            c.font = fnt
            c.alignment = Alignment(horizontal=ha, vertical="center",
                                     indent=1 if ha == "left" else 0)

    # 淨效益條件顏色
    net_c = ws["B9"]
    net_c.font = Font(bold=True, size=11,
                       color="059669" if (net_c.value or 0) >= 0 else "DC2626")
    ws["B7"].number_format = "0.0%"
    ws["B8"].number_format = "$#,##0"
    ws["B9"].number_format = "$#,##0"
    ws.row_dimensions[10].height = 5    # 細分隔

    # 績效指標 section 標題 (row 11)
    ws.merge_cells("A11:C11")
    ws["A11"].fill = F_SECTION
    ws["A11"].font = Font(color="FFFFFF", bold=True, size=11)
    ws["A11"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[11].height = 22

    # 欄位標題 (row 12)
    for col in range(1, 4):
        c = ws.cell(row=12, column=col)
        c.fill = F_HDR; c.border = FULL_B
        c.font = Font(color="FFFFFF", bold=True, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[12].height = 20

    # 績效資料 (rows 13–23)，交替底色
    for r in range(13, 24):
        f = F_ALT if r % 2 == 0 else F_WHITE
        ws.row_dimensions[r].height = 18
        for col, (ha, fnt) in enumerate([
            ("left",  Font(size=10)),
            ("right", Font(bold=True, size=11)),
            ("left",  Font(color="6B7280", size=9, italic=True)),
        ], 1):
            c = ws.cell(row=r, column=col)
            c.fill = f; c.border = FULL_B
            c.font = fnt
            c.alignment = Alignment(horizontal=ha, vertical="center",
                                     indent=1 if ha == "left" else 0)

    # 數字格式（績效）
    for row_n, fmt in [(17,"0.000"),(18,"0.000"),(19,"0.000"),(20,"0.000"),
                       (21,"$#,##0"),(22,"$#,##0"),(23,"0.00")]:
        ws.cell(row=row_n, column=2).number_format = fmt
    ws["B23"].font = Font(bold=True, size=11, color="D97706")

    auto_w(ws)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 32

    ws.print_area = "A1:E23"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 9
    ws.oddHeader.center.text = "風控管理報表"
    ws.oddFooter.right.text = "第 &P 頁，共 &N 頁"

    # ══ Sheets 2 & 3 ════════════════════════════════════
    sheets_cfg = [
        ("高風險交易",   "F59E0B", False,
         ["交易ID","時間","金額 (USD)","風險評分","類型","建議行動"],
         highrisk_export_rows),
        ("完整交易明細", "475569", True,
         ["Transaction_ID","時間","金額_USD","風險評分","風險等級",
          "真實標籤","系統判定","結果","處置優先級"],
         detail_rows),
    ]

    for tbl_idx, (title, tab_clr, large, hdrs, rows) in enumerate(sheets_cfg, 1):
        sh = wb.create_sheet(title)
        sh.sheet_view.showGridLines = False
        sh.sheet_properties.tabColor = tab_clr
        sh.freeze_panes = "B2"

        sh.append(hdrs)
        for row in rows: sh.append(row)
        max_r = sh.max_row

        # 欄位標題
        for col_idx in range(1, len(hdrs) + 1):
            c = sh.cell(row=1, column=col_idx)
            c.fill = F_HDR; c.border = FULL_B
            c.font = Font(color="FFFFFF", bold=True, size=10)
            c.alignment = Alignment(horizontal="center", vertical="center")
        sh.row_dimensions[1].height = 22

        # 行高 & 框線（小表才逐格）
        for row_idx in range(2, max_r + 1):
            sh.row_dimensions[row_idx].height = 18
            if not large:
                for col_idx in range(1, len(hdrs) + 1):
                    sh.cell(row=row_idx, column=col_idx).border = FULL_B

        # Excel Table
        if max_r > 1:
            tbl = Table(displayName=f"ReportTable{tbl_idx}",
                        ref=f"A1:{get_column_letter(len(hdrs))}{max_r}")
            tbl.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2", showRowStripes=not large)
            sh.add_table(tbl)

        all_range = f"A2:{get_column_letter(len(hdrs))}{max_r}"

        # 條件格式：TP/FP/FN/TN 整行底色
        result_key = "結果" if "結果" in hdrs else "類型"
        if result_key in hdrs:
            rc = get_column_letter(hdrs.index(result_key) + 1)
            sh.conditional_formatting.add(all_range,
                FormulaRule(formula=[f'ISNUMBER(SEARCH("FN",${rc}2))'], fill=F_RED))
            sh.conditional_formatting.add(all_range,
                FormulaRule(formula=[f'ISNUMBER(SEARCH("FP",${rc}2))'], fill=F_ORANGE))
            if large:
                sh.conditional_formatting.add(all_range,
                    FormulaRule(formula=[f'OR(${rc}2="TP",${rc}2="TN")'], fill=F_GREEN))

        # 條件格式：風險等級底色
        if "風險等級" in hdrs:
            rlc = get_column_letter(hdrs.index("風險等級") + 1)
            rl_range = f"{rlc}2:{rlc}{max_r}"
            sh.conditional_formatting.add(rl_range,
                FormulaRule(formula=[f'${rlc}2="高"'],
                            fill=F_RED,    font=Font(color="DC2626", bold=True)))
            sh.conditional_formatting.add(rl_range,
                FormulaRule(formula=[f'${rlc}2="中"'],
                            fill=F_ORANGE, font=Font(color="D97706", bold=True)))
            sh.conditional_formatting.add(rl_range,
                FormulaRule(formula=[f'${rlc}2="低"'],
                            fill=F_GREEN,  font=Font(color="059669", bold=True)))

        # Data Bar：風險評分
        if "風險評分" in hdrs:
            rsc = get_column_letter(hdrs.index("風險評分") + 1)
            sh.conditional_formatting.add(f"{rsc}2:{rsc}{max_r}",
                DataBarRule(start_type="num", start_value=0,
                            end_type="num",   end_value=1,
                            color="1E9FFB",   showValue=True,
                            minLength=0, maxLength=100))

        # 數字格式
        for row_idx in range(2, max_r + 1):
            if "風險評分" in hdrs:
                sh.cell(row=row_idx,
                        column=hdrs.index("風險評分")+1).number_format = "0.0000"
            for amt_col in ("金額 (USD)", "金額_USD"):
                if amt_col in hdrs:
                    sh.cell(row=row_idx,
                            column=hdrs.index(amt_col)+1).number_format = "$#,##0.00"

        auto_w(sh)
        sh.page_setup.orientation = "landscape"
        sh.page_setup.paperSize = 9
        sh.page_setup.fitToWidth = 1
        sh.oddHeader.center.text = title
        sh.oddFooter.right.text = "第 &P 頁，共 &N 頁"

    output = io.BytesIO()
    wb.save(output)
    xlsx_content = base64.b64encode(output.getvalue()).decode("ascii")
    store = {
        "filename": "daily_report.xlsx",
        "content": xlsx_content,
        "base64": True,
        "type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    return content, store, {"display": "inline-block"}


@app.callback(
    Output("report-download", "data"),
    Input("download-btn",     "n_clicks"),
    State("report-store",     "data"),
    prevent_initial_call=True,
)
def download_report(n_clicks, store_data):
    if not store_data:
        return no_update
    return {
        "content": store_data["content"],
        "filename": store_data["filename"],
        "base64": store_data.get("base64", False),
        "type": store_data.get("type", "text/csv"),
    }


@app.callback(
    Output("tx-detail-panel", "children"),
    Input("stream-table",     "selected_rows"),
    State("stream-table",     "data"),
    State("settings-store",   "data"),
    prevent_initial_call=True,
)
def show_tx_detail(selected_rows, data, settings):
    if not selected_rows or not data:
        return None
    row = data[selected_rows[0]]
    if row.get("交易ID") == "-" or "_risk_raw" not in row:
        return None

    risk        = row["_risk_raw"]
    sop_text    = row["_sop"]
    true_label  = row.get("_true_label", -1)
    amount_eur  = row.get("_amount_eur", 0.0)
    time_sec    = row.get("_time_sec",   0.0)
    threshold   = settings.get("threshold", 0.5)

    # 判斷 TP/TN/FP/FN（以當前閥值即時計算）
    predicted = risk >= threshold
    if   true_label == 1 and predicted:      verdict = "TP — 成功攔截";       verdict_color = COLORS["safe"]
    elif true_label == 0 and not predicted:  verdict = "TN — 正確放行";       verdict_color = COLORS["safe"]
    elif true_label == 0 and predicted:      verdict = "FP — 誤報（實際正常）"; verdict_color = COLORS["warning"]
    elif true_label == 1 and not predicted:  verdict = "FN — 漏抓（實際詐騙）"; verdict_color = COLORS["fraud"]
    else:                                    verdict = "—";                   verdict_color = COLORS["text_muted"]

    true_label_txt   = "詐騙" if true_label == 1 else "正常" if true_label == 0 else "—"
    true_label_color = COLORS["fraud"] if true_label == 1 else COLORS["safe"]
    status_color     = COLORS["fraud"] if predicted else COLORS["safe"]
    bar_color        = COLORS["fraud"] if risk > 0.7 else COLORS["warning"] if risk > 0.3 else COLORS["safe"]

    # 時間轉換：秒 → 小時分鐘
    h = int(time_sec // 3600)
    m = int((time_sec % 3600) // 60)
    time_display = f"T+{h}h {m:02d}m"

    def info_row(label, value, value_color=COLORS["text"]):
        return html.Div([
            html.Span(label, style={"color": COLORS["text_muted"], "fontSize": "11px",
                                    "width": "90px", "display": "inline-block"}),
            html.Span(value, style={"color": value_color, "fontFamily": "monospace",
                                    "fontSize": "12px", "fontWeight": "bold"}),
        ], style={"marginBottom": "5px"})

    return dbc.Card([
        dbc.CardBody([
            # Header
            html.Div([
                html.Span(row["交易ID"], style={
                    "fontFamily": "monospace", "fontSize": "14px",
                    "color": COLORS["cyan"], "fontWeight": "bold"
                }),
                html.Span(row["系統判定狀態"], style={
                    "fontSize": "12px", "color": status_color,
                    "fontWeight": "bold", "marginLeft": "14px"
                }),
            ], style={"marginBottom": "10px"}),

            # Risk bar
            html.Div([
                html.Div(style={"height": "4px", "borderRadius": "2px",
                                "backgroundColor": COLORS["border"]}),
                html.Div(style={"height": "4px", "borderRadius": "2px",
                                "width": f"{min(risk*100, 100):.1f}%",
                                "backgroundColor": bar_color, "marginTop": "-4px"}),
            ], style={"marginBottom": "10px"}),

            # Info grid
            info_row("風險評分",  f"{risk:.4f}  ({risk*100:.1f}%)", bar_color),
            info_row("交易金額",  f"${amount_eur:,.2f}"),
            info_row("發生時間",  time_display),
            html.Hr(style={"borderColor": COLORS["border"], "margin": "8px 0"}),
            info_row("真實標籤",  true_label_txt,  true_label_color),
            info_row("模型結果",  verdict,          verdict_color),
            info_row("處置優先",  row["處置優先級"], COLORS["text_muted"]),
            html.Hr(style={"borderColor": COLORS["border"], "margin": "8px 0"}),
            html.Pre(sop_text, style={
                "color": COLORS["text_muted"], "fontSize": "11px",
                "whiteSpace": "pre-wrap", "marginBottom": 0,
                "fontFamily": "monospace", "lineHeight": "1.6",
            }),
        ], style={"padding": "12px"})
    ], style={"backgroundColor": COLORS["bg_input"],
              "border": f"1px solid {COLORS['border']}"})


if __name__ == "__main__":
    print("\n" + "="*50)
    print("請開啟瀏覽器前往：http://127.0.0.1:8050")
    print("="*50 + "\n")
    app.run(debug=False)
